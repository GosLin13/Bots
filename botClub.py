import logging
from aiogram import Bot, Dispatcher, types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import Command
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
import asyncio
from openpyxl import Workbook, load_workbook

# Токен бота
BOT_TOKEN = ""
bot = Bot(token=BOT_TOKEN)

# Создаем объект MemoryStorage для хранения состояний
storage = MemoryStorage()

# Создаем Dispatcher с хранилищем состояний
dp = Dispatcher(storage=storage)

# Создаем Excel файл и таблицу, если она не существует
file_name = "privazka.xlsx"
try:
    wb = load_workbook(file_name)
    sheet = wb.active
except FileNotFoundError:
    wb = Workbook()
    sheet = wb.active
    sheet.append(["adult_id", "adult_username", "child_username"])  # Заголовки
    wb.save(file_name)

# Запись данных о взрослом в таблицу
def write_adult_to_excel(adult_id, adult_username, file_name="privazka.xlsx"):
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
        sheet.append([adult_id, adult_username, ""])  # Записываем данные о взрослом
        wb.save(file_name)
    except Exception as e:
        print(f"Ошибка при записи в Excel: {e}")


# Стартовое сообщение
@dp.message(Command("start"))
async def start_handler(message: types.Message):
    # Клавиатура с двумя кнопками
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=">18", callback_data="over_18")],
            [InlineKeyboardButton(text="<18", callback_data="under_18")]
        ]
    )
    # Отправляем сообщение с клавиатурой
    await message.answer("Привет! Пожалуйста, выбери свою возрастную категорию:", reply_markup=keyboard)


# Обработка нажатия кнопок для выбора возраста
@dp.callback_query(lambda c: c.data in ["over_18", "under_18"])
async def age_category_handler(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    selected_option = callback_query.data

    if selected_option == "over_18":
        # Записываем adult_id и username в таблицу
        adult_username = callback_query.from_user.username
        write_adult_to_excel(adult_id=user_id, adult_username=adult_username)

        # Показываем меню взрослого
        adult_menu = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="Привязать аккаунт <18", callback_data="bind_child")]
            ]
        )
        await callback_query.message.answer("Вы выбрали возраст >18. Вы можете привязать аккаунт ребенка.",
                                            reply_markup=adult_menu)
        # Сохраняем информацию о том, что это взрослый
        await state.update_data(role="adult")

    elif selected_option == "under_18":
        # Сообщаем, что это для ребенка
        await callback_query.message.answer("Вы выбрали категорию <18. Вы не можете делать оплату.")

    # Добавим явное подтверждение обработки callback
    await callback_query.answer()

# Определяем состояние для FSM
class BindChildState(StatesGroup):
    waiting_for_child_username = State()

# Обработка кнопки для привязки аккаунта ребенка
@dp.callback_query(lambda c: c.data == "bind_child")
async def bind_child_handler(callback_query: types.CallbackQuery, state: FSMContext):
    # Запрашиваем у взрослого @username ребенка для привязки
    await callback_query.message.answer("Пожалуйста, впишите @username ребенка для привязки.")
    # Устанавливаем состояние ожидания ввода ребенка
    await state.set_state(BindChildState.waiting_for_child_username)

    # Сохраняем ID взрослого для привязки
    adult_id = callback_query.from_user.id
    await state.update_data(adult_id=adult_id)

    # Добавляем явное подтверждение
    await callback_query.answer()

# Обработка ввода username ребенка
@dp.message(BindChildState.waiting_for_child_username)
async def process_child_username(message: types.Message, state: FSMContext):
    # Получаем введенный username
    child_username = message.text

    # Проверяем, что это username в формате @username
    if not child_username.startswith("@"):
        await message.answer("Пожалуйста, введите корректный username (должен начинаться с @).")
        return

    # Убираем символ @ для сохранения в таблице
    child_username = child_username[1:]

    # Получаем ID взрослого из состояния
    user_data = await state.get_data()
    adult_id = user_data["adult_id"]

    # Записываем username ребенка в таблицу
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            if row[0].value == adult_id:
                row[2].value = child_username  # Записываем username ребенка в соответствующую строку
                wb.save(file_name)
                break
        else:
            await message.answer("Ошибка: взрослый не найден в таблице.")
            await state.clear()
            return

        await message.answer(f"Аккаунт ребенка @{child_username} успешно привязан к вашему аккаунту.")
    except Exception as e:
        await message.answer(f"Ошибка при записи в таблицу: {e}")

    # Завершаем состояние
    await state.clear()

# Запуск бота
async def main():
    logging.basicConfig(level=logging.INFO)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
